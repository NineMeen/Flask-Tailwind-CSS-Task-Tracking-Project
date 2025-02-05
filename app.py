from gevent import monkey
monkey.patch_all()

# Now import other modules
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, render_template_string, current_app, abort
from flask_login import LoginManager, login_required, current_user, login_user, logout_user
from flask_socketio import SocketIO, emit, join_room, leave_room
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import os
from datetime import datetime, timedelta
from models import db, User, Team, Migration, MigrationFile, MigrationLog, Notification
from ldap3 import Server, Connection, ALL, NTLM
from flask_migrate import Migrate
from sqlalchemy.pool import NullPool
from flask_wtf import FlaskForm
from flask_mail import Mail, Message
from jinja2 import Template
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
from pytz import timezone

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key'  # Change this in production
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///migration_tracker.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'poolclass': NullPool,  # Disable connection pooling
    'pool_pre_ping': True,  # Optional: Enable connection health checks
}
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Mail configuration
app.config['MAIL_SERVER'] = '10.10.10.160'  # Change according to your email provider
app.config['MAIL_PORT'] = 25
app.config['MAIL_USE_TLS'] = True
# app.config['MAIL_USERNAME'] = 'your-email@gmail.com'  # Change this
# app.config['MAIL_PASSWORD'] = 'your-app-password'  # Change this
app.config['MAIL_DEFAULT_SENDER'] = 'no-reply-worktracking@symphony.net.th'  # Change this

mail = Mail(app)


socketio = SocketIO(app)
# Initialize extensions
db.init_app(app)
migrate = Migrate(app, db)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# LDAP Configuration
LDAP_SERVER = '10.10.11.45'
LDAP_PORT = 389
LDAP_BASE_DN = 'OU=WorkTracking,DC=test,DC=net,DC=th'

# Initialize SocketIO
socketio = SocketIO(app, async_mode='gevent')

# At the top of your app.py
TIMEZONE = timezone('Asia/Bangkok')  # Set to your timezone

@socketio.on('disconnect')
def handle_disconnect():
    db.session.remove()

def authenticate_ldap(username, password):
    try:
        server = Server(
            LDAP_SERVER,
            port=LDAP_PORT,
            use_ssl=False,
            get_info=ALL
        )
        
        user_dn = username
        conn = Connection(
            server,
            user=user_dn,
            password=password,
            authentication='SIMPLE'
        )
        
        if conn.bind():
            conn.search(
                LDAP_BASE_DN,
                f'(&(objectClass=person)(sAMAccountName={username}))',
                attributes=['displayName', 'mail', 'memberOf']
            )
            
            if len(conn.entries) > 0:
                user_data = conn.entries[0]
                print(f"LDAP User Data: {user_data}")  # Debug print
                
                # Extract group names from memberOf DNs
                groups = []
                if hasattr(user_data, 'memberOf'):
                    for group_dn in user_data.memberOf:
                        try:
                            print(f"Processing group DN: {group_dn}")  # Debug print
                            cn_parts = [part for part in group_dn.split(',') if part.startswith('CN=')]
                            if cn_parts:
                                group_name = cn_parts[0].replace('CN=', '').strip()
                                groups.append(group_name)
                        except Exception as e:
                            print(f"Error processing group: {e}")  # Debug print
                            continue
                
                print(f"Extracted groups: {groups}")  # Debug print
                
                return {
                    'status': 'success',
                    'message': 'Authentication successful',
                    'user': {
                        'displayName': user_data.displayName.value if hasattr(user_data, 'displayName') else None,
                        'email': user_data.mail.value if hasattr(user_data, 'mail') else None,
                        'memberOf': groups
                    }
                }
            return {
                'status': 'error',
                'message': 'Authentication successful but no user data found'
            }
        else:
            return {
                'status': 'error',
                'message': 'Invalid credentials'
            }
            
    except Exception as e:
        print(f"LDAP Authentication Error: {e}")  # Debug print
        return {
            'status': 'error',
            'message': str(e)
        }

@login_manager.user_loader
def load_user(user_id):
    if user_id is None:
        return None
    try:
        # Try to load user by numeric ID first
        return User.query.get(int(user_id))
    except ValueError:
        # If user_id is not numeric, try to find user by username
        return User.query.filter_by(username=user_id).first()

def log_action(migration_id, user_id, action, details=None):
    log = MigrationLog(
        migration_id=migration_id,
        user_id=user_id,
        action=action,
        details=details
    )
    db.session.add(log)
    db.session.commit()

def send_notification_email(recipients, subject, template_name, **kwargs):
    """
    Generic function to send HTML emails using Flask's template rendering
    """
    try:
        # Load template using Flask's render_template
        template_path = f'email/{template_name}.html'
        html_content = render_template(template_path, **kwargs)
        
        msg = Message(
            subject=subject,
            recipients=recipients,
            html=html_content
        )
        mail.send(msg)
        print(f"Email sent successfully with template: {template_path}")  # Debug print
        return True
    except Exception as e:
        print(f"Error sending email: {str(e)}")  # Debug print
        print(f"Template kwargs: {kwargs}")  # Debug print
        return False

@app.route('/')
@login_required
def index():
    if current_user.team.name == 'SD':
        migrations = Migration.query.filter_by(created_by=current_user.id).filter_by(is_deleted=False).all()
    elif current_user.team.name in ['SA', 'NS']:
        migrations = Migration.query.filter(
            (Migration.status != 'completed') & 
            ((Migration.assigned_to == current_user.id) | 
             (Migration.assigned_to == None))
        ).filter_by(is_deleted=False).all()
    else:  # Admin
        migrations = Migration.query.filter_by(is_deleted=False).all()
    
    return render_template('task/task.html', migrations=migrations ,User=User)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
        
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # Try LDAP authentication
        ldap_result = authenticate_ldap(username, password)
        
        if ldap_result['status'] == 'success':
            # Check if user exists in local database
            user = User.query.filter_by(username=username).first()
            
            if user is not None and not user.is_active:
                flash('Your account is disabled. Please contact an administrator.', 'error')
                return redirect(url_for('login'))
                
            if user is None:
                # Get team from LDAP memberOf attribute
                ldap_groups = ldap_result['user']['memberOf']
                team_name = None
                
                # Debug print
                print(f"LDAP Groups: {ldap_groups}")
                
                # Handle if memberOf is a list or string
                if isinstance(ldap_groups, list) and ldap_groups:
                    # Try to find the first valid team
                    for group in ldap_groups:
                        if 'GWTK01' in group:
                            team_name = 'SA'
                            break
                        elif 'GWTK02' in group:
                            team_name = 'SD'
                            break
                        elif 'GWTK03' in group:
                            team_name = 'NS'
                            break
                elif isinstance(ldap_groups, str):
                    # Single group
                    if 'GWTK01' in ldap_groups:
                        team_name = 'SA'
                    elif 'GWTK02' in ldap_groups:
                        team_name = 'SD'
                    elif 'GWTK03' in ldap_groups:
                        team_name = 'NS'
                
                # If no valid team found, set default
                if not team_name:
                    print(f"No valid team found in groups: {ldap_groups}")
                    team_name = 'NS'  # Default team
                
                print(f"Assigned team: {team_name}")  # Debug print
                
                # Create new user with team from LDAP
                user = User(
                    username=username,
                    email=ldap_result['user']['email'] if ldap_result['user'].get('email') else f"{username}@symphony.net.th",
                    password=generate_password_hash(password),
                    team_id=get_team_id(team_name),
                    is_admin='GWTKADMIN' in str(ldap_groups),
                    is_active=True
                )
                try:
                    db.session.add(user)
                    db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    print(f"Error creating user: {e}")  # Debug print
                    flash('Error creating user account', 'error')
                    return redirect(url_for('login'))
            
            # Log the user in
            login_user(user)
            next_page = request.args.get('next')
            return redirect(next_page if next_page else url_for('index'))
            
        # Handle LDAP connection errors specifically
        elif 'socket connection error' in ldap_result['message'].lower():
            flash('Unable to connect to authentication server. Please try again later or contact your administrator.', 'warning')
            # Log the error for administrators
            print(f"LDAP Connection Error: {ldap_result['message']}")
            return render_template('login.html', offline_mode=True)
        else:
            flash('Invalid username or password', 'error')
            
    return render_template('login.html', offline_mode=False)

def get_team_id(team_name):
    """Helper function to map team names to local team IDs"""
    team_mapping = {
        'SA': 1,
        'SD': 2,
        'NS': 3,
        'ADMIN': 1  # ADMIN users are part of SA team
    }
    return team_mapping.get(team_name, 3)  # Default to NS team if unknown

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

# @app.route('/admin')
# @login_required
# def admin_dashboard():
#     return render_template('admin/index.html')


@app.route('/migration/new', methods=['GET', 'POST'])
@login_required
def create_migration():
    if current_user.team.name != 'SD' and not current_user.is_admin:
        flash('Only SD team members can create migration requests.', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        scheduled_date_str = request.form.get('scheduled_date')
        scheduled_time_str = request.form.get('scheduled_time')
        
        if scheduled_date_str and scheduled_time_str:
            try:
                scheduled_datetime = datetime.strptime(
                    f"{scheduled_date_str} {scheduled_time_str}",
                    "%Y-%m-%d %H:%M"
                )
            except ValueError:
                flash('Invalid date or time format', 'error')
                return redirect(url_for('create_migration'))
        else:
            scheduled_datetime = None

        # Create new migration with current timestamp
        migration = Migration(
            title=request.form['title'],
            description=request.form['description'],
            customer_name=request.form['customer_name'],
            customer_contact=request.form['customer_contact'],
            created_by=current_user.id,
            status='waiting',
            scheduled_date=scheduled_datetime,
            created_at=datetime.now(TIMEZONE)  # Add current timestamp
        )
        
        db.session.add(migration)
        db.session.commit()

        # Create unique folder for this migration
        migration_folder = os.path.join(app.config['UPLOAD_FOLDER'], f'migration_{migration.id}')
        os.makedirs(migration_folder, exist_ok=True)

        # Handle multiple file uploads
        files = request.files.getlist('files[]')
        for file in files:
            if file and file.filename:
                filename = secure_filename(file.filename)
                file_path = os.path.join(migration_folder, filename)
                file.save(file_path)
                
                migration_file = MigrationFile(
                    migration_id=migration.id,
                    filename=filename,
                    file_path=file_path,
                    file_type='attachment',
                    uploaded_at=datetime.now(TIMEZONE)  # Add timestamp for files too
                )
                db.session.add(migration_file)

        
        db.session.commit()
        log_action(migration.id, current_user.id, 'created')

        # Create notification for SA team members
        sa_team = Team.query.filter_by(name='SA').first()
        if sa_team:
            sa_users = User.query.filter_by(team_id=sa_team.id, is_active=True).all()
            for sa_user in sa_users:
                notification = Notification(
                    user_id=sa_user.id,
                    message=f'New migration request: {migration.title}',
                    migration_id=migration.id,
                    created_at=datetime.now(TIMEZONE)
                )
                db.session.add(notification)
            
            try:
                db.session.commit()
                # Emit socket event for real-time notification
                socketio.emit('new_notification', {
                    'message': f'New migration request: {migration.title}',
                    'migration_id': migration.id,
                    'created_at': datetime.now(TIMEZONE).strftime('%Y-%m-%d %H:%M:%S')
                }, room='sa_team')
                
            except Exception as e:
                print(f"Error creating notifications: {str(e)}")
                db.session.rollback()

        flash('Migration request created successfully.', 'success')
        return redirect(url_for('index'))

    return render_template('migration/create.html')

@app.route('/migration/<int:id>', methods=['GET'])
@login_required
def view_migration(id):
    migration = Migration.query.get_or_404(id)
    migration_files = MigrationFile.query.filter_by(migration_id=migration.id).all()
    return render_template('migration/view.html', 
                         migration=migration, 
                         migration_files=migration_files,
                         User=User)

@app.route('/migration/<int:id>/acknowledge', methods=['POST'])
@login_required
def acknowledge_migration(id):
    if current_user.team.name != 'SA':
        flash('Only SA team members can acknowledge migrations.', 'error')
        return redirect(url_for('index'))

    migration = Migration.query.get_or_404(id)
    migration.status = 'acknowledged'
    migration.acknowledged_at = datetime.now(TIMEZONE)
    db.session.commit()


    log_action(migration.id, current_user.id, 'acknowledged')
    flash('Migration acknowledged successfully.', 'success')
    return redirect(url_for('index'))

@app.route('/migration/<int:id>/assign', methods=['POST'])
@login_required
def assign_migration(id):
    if current_user.team.name != 'SA':
        flash('Only SA team members can assign migrations.', 'error')
        return redirect(url_for('index'))

    migration = Migration.query.get_or_404(id)
    migration.assigned_to = current_user.id
    migration.status = 'in_progress'
    db.session.commit()

    log_action(migration.id, current_user.id, 'assigned', f'Assigned to {current_user.username}')
    flash('Migration assigned successfully.', 'success')
    return redirect(url_for('index'))

@app.route('/migration/<int:id>/update-status', methods=['POST'])
@login_required
def update_migration_status(id):
    if current_user.team.name != 'SA':
        flash('Only SA team members can update migration status.', 'error')
        return redirect(url_for('index'))

    migration = Migration.query.get_or_404(id)
    new_status = request.form.get('status')
    
    if new_status in ['completed', 'rollback', 'in_progress']:
        migration.status = new_status
        migration.completed_at = datetime.now(TIMEZONE)
        
        # Handle result file upload
        if 'result_file' in request.files:
            file = request.files['result_file']
            if file and file.filename:
                filename = secure_filename(file.filename)
                migration_folder = os.path.join(app.config['UPLOAD_FOLDER'], f'migration_{migration.id}')
                file_path = os.path.join(migration_folder, filename)
                file.save(file_path)
                
                migration_file = MigrationFile(
                    migration_id=migration.id,
                    filename=filename,
                    file_path=file_path,
                    file_type='result'
                )
                db.session.add(migration_file)
        
        db.session.commit()
        log_action(migration.id, current_user.id, f'status_updated', f'Status changed to {new_status}')
        
        # Send email notification to SD team member
        if new_status in ['completed', 'rollback']:
            creator = User.query.get(migration.created_by)
            if creator and creator.email:
                recipients = [creator.email]
                # Add test emails one by one
                recipients.append('thaktechin.bo.64@ubu.ac.th')
                recipients.append('patthamakm54@gmail.com')
                
                template_data = {
                    'migration': migration,
                    'updater': current_user,
                    'new_status': new_status,
                    'view_url': url_for('view_migration', id=migration.id, _external=True)
                }
                
                send_notification_email(
                    recipients=recipients,
                    subject=f'Migration Status Update: {migration.title}',
                    template_name='status_update',
                    **template_data
                )
                print(f"Sending email to: {recipients}")  # Debug print
        flash(f'Migration status updated to {new_status}', 'success')
    
    return redirect(url_for('view_migration', id=id))


@app.route('/file/<int:file_id>/view')
@login_required
def view_file(file_id):
    file = MigrationFile.query.get_or_404(file_id)
    return send_file(file.file_path, as_attachment=False)

@app.route('/file/<int:file_id>/download')
@login_required
def download_file(file_id):
    file = MigrationFile.query.get_or_404(file_id)
    return send_file(file.file_path, as_attachment=True, download_name=file.filename)

@app.route('/team')
@login_required
def team_page():
    if current_user.team.name == 'SA' or current_user.team.name == 'SD':
        team_members = User.query.filter_by(team_id=current_user.team_id).all()
        
        # Get selected member from query parameter
        selected_member_id = request.args.get('member_id', type=int)
        
        # Get migrations based on selection
        if selected_member_id:
            if current_user.team.name == 'SA':
                migrations = Migration.query.filter(
                    Migration.status.in_(['in_progress', 'waiting', 'acknowledged']),
                    Migration.assigned_to == selected_member_id
                ).filter_by(is_deleted=False).all()
            else:  # SD team
                migrations = Migration.query.filter(
                    Migration.status.in_(['in_progress', 'waiting', 'acknowledged']),
                    Migration.created_by == selected_member_id
                ).filter_by(is_deleted=False).all()
        else:
            # Show all active tasks if no member selected
            if current_user.team.name == 'SA':
                migrations = Migration.query.filter(
                    Migration.status.in_(['in_progress', 'waiting', 'acknowledged'])
                ).filter_by(is_deleted=False).all()
            else:  # SD team
                migrations = Migration.query.filter(
                    Migration.status.in_(['in_progress', 'waiting', 'acknowledged']),
                    Migration.created_by.in_([member.id for member in team_members])
                ).filter_by(is_deleted=False).all()
        
        return render_template('task/team.html', 
                             team_members=team_members, 
                             migrations=migrations,
                             selected_member_id=selected_member_id,
                             User=User)
    else:
        flash('Access denied. Only SA and SD team members can view this page.', 'error')
        return redirect(url_for('index'))

# Initialize the database
with app.app_context():
    db.create_all()
    
    # Create initial teams if they don't exist
    teams = ['SD', 'SA', 'NS']
    for team_name in teams:
        if not Team.query.filter_by(name=team_name).first():
            team = Team(name=team_name)
            db.session.add(team)
    
    # Create admin user if it doesn't exist
    if not User.query.filter_by(username='admin').first():
        admin = User(
            username='admin',
            password=generate_password_hash('admin'),
            email='admin@example.com',
            team_id=1,
            is_admin=True
        )
        db.session.add(admin)
    
    db.session.commit()

@app.route('/admin/create-user', methods=['GET', 'POST'])
@login_required
def create_user():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form['username']
        email = request.form['email']
        password = request.form['password']
        confirm_password = request.form['confirm_password']
        team_id = request.form['team_id']
        is_admin = 'is_admin' in request.form

        # Basic validation
        if password != confirm_password:
            flash('Passwords do not match', 'error')
            return redirect(url_for('create_user'))

        if User.query.filter_by(username=username).first():
            flash('Username already exists', 'error')
            return redirect(url_for('create_user'))

        if User.query.filter_by(email=email).first():
            flash('Email already exists', 'error')
            return redirect(url_for('create_user'))

        # Create new user
        new_user = User(
            username=username,
            email=email,
            password=generate_password_hash(password),
            team_id=team_id,
            is_admin=is_admin
        )
        
        try:
            db.session.add(new_user)
            db.session.commit()
            flash('User created successfully', 'success')
            return redirect(url_for('create_user'))
        except Exception as e:
            db.session.rollback()
            flash('Error creating user', 'error')
            return redirect(url_for('create_user'))

    # GET request - display form
    teams = Team.query.all()
    return render_template('admin/create.html', teams=teams)

@app.route('/search')
@login_required
def search():
    # Get search parameters
    search_query = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    page = request.args.get('page', 1, type=int)
    per_page = 10  # Number of items per page
    
    # Base query
    query = Migration.query.filter_by(is_deleted=False)

    # Apply status filter if specified
    if status_filter:
        query = query.filter(Migration.status == status_filter)
    else:
        # Show all tasks when searching, but only completed and rollback by default
        if search_query:
            query = query.filter(Migration.status.in_(['completed', 'rollback', 'in_progress', 'waiting']))
        else:
            query = query.filter(Migration.status.in_(['completed', 'rollback']))

    # Apply search if specified
    if search_query:
        search = f"%{search_query}%"
        query = query.filter(
            db.or_(
                Migration.title.ilike(search),
                Migration.customer_name.ilike(search)
            )
        )

    # Order by completion date and paginate
    migrations = query.order_by(Migration.completed_at.desc()).paginate(
        page=page,
        per_page=per_page,
        error_out=False
    )
    
    return render_template('search.html', migrations=migrations)

@app.route('/notifications')
@login_required
def get_notifications():
    notifications = Notification.query.filter_by(
        user_id=current_user.id,
        read=False
    ).order_by(Notification.created_at.desc()).all()
    
    return jsonify([{
        'id': n.id,
        'message': n.message,
        'migration_id': n.migration_id,
        'created_at': n.created_at.strftime('%Y-%m-%d %H:%M:%S')
    } for n in notifications])

@app.route('/notifications/mark-read/<int:notification_id>', methods=['POST'])
@login_required
def mark_notification_read(notification_id):
    notification = Notification.query.get_or_404(notification_id)
    if notification.user_id == current_user.id:
        notification.read = True
        db.session.commit()
    return jsonify({'success': True})

@app.route('/notifications/mark-all-read', methods=['POST'])
@login_required
def mark_all_notifications_read():
    notifications = Notification.query.filter_by(
        user_id=current_user.id,
        read=False
    ).all()
    
    for notification in notifications:
        notification.read = True
    
    db.session.commit()
    return jsonify({'success': True})

# Socket.IO event handlers
@socketio.on('connect')
def handle_connect():
    if current_user.is_authenticated:
        if current_user.team.name == 'SA':
            join_room('sa_team')
            print(f"User {current_user.username} joined SA team room")

@socketio.on('disconnect')
def handle_disconnect():
    if current_user.is_authenticated:
        if current_user.team.name == 'SA':
            leave_room('sa_team')
            print(f"User {current_user.username} left SA team room")

@app.route('/migration/<int:id>/edit-page')
@login_required
def edit_migration_page(id):
    if current_user.team.name != 'SD' and not current_user.is_admin:
        flash('Only SD team members can edit migrations.', 'error')
        return redirect(url_for('index'))
    
    migration = Migration.query.get_or_404(id)
    if migration.created_by != current_user.id and not current_user.is_admin:
        flash('You can only edit your own migrations.', 'error')
        return redirect(url_for('index'))
    
    # Get the migration files
    migration_files = MigrationFile.query.filter_by(migration_id=id).all()
    form = FlaskForm()
    return render_template('migration/edit.html', 
                         migration=migration,
                         migration_files=migration_files,
                         form=form)  # Pass files and form to template

@app.route('/migration/<int:id>/edit', methods=['POST'])
@login_required
def edit_migration(id):
    if current_user.team.name != 'SD' and not current_user.is_admin:
        flash('Only SD team members can edit migrations.', 'error')
        return redirect(url_for('index'))

    migration = Migration.query.get_or_404(id)
    
    # Check if the user is the creator or admin
    if migration.created_by != current_user.id and not current_user.is_admin:
        flash('You can only edit your own migrations.', 'error')
        return redirect(url_for('index'))
    
    # Update scheduled date if provided
    scheduled_date_str = request.form.get('scheduled_date')
    scheduled_time_str = request.form.get('scheduled_time')
    
    if scheduled_date_str and scheduled_time_str:
        try:
            scheduled_datetime = datetime.strptime(
                f"{scheduled_date_str} {scheduled_time_str}",
                "%Y-%m-%d %H:%M"
            )
            migration.scheduled_date = scheduled_datetime
        except ValueError:
            flash('Invalid date or time format', 'error')
            return redirect(url_for('edit_migration_page', id=id))
    else:
        migration.scheduled_date = None
    
    # Update other migration details
    migration.title = request.form['title']
    migration.description = request.form['description']
    migration.customer_name = request.form['customer_name']
    migration.customer_contact = request.form['customer_contact']
    
    # Handle file uploads
    files = request.files.getlist('files[]')
    if files:
        migration_folder = os.path.join(app.config['UPLOAD_FOLDER'], f'migration_{migration.id}')
        os.makedirs(migration_folder, exist_ok=True)
        
        for file in files:
            if file and file.filename:
                filename = secure_filename(file.filename)
                file_path = os.path.join(migration_folder, filename)
                file.save(file_path)
                
                migration_file = MigrationFile(
                    migration_id=migration.id,
                    filename=filename,
                    file_path=file_path,
                    file_type='attachment'
                )
                db.session.add(migration_file)

    try:
        db.session.commit()
        log_action(migration.id, current_user.id, 'edited', 'Migration details updated')
        flash('Migration updated successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error updating migration.', 'error')
    
    return redirect(url_for('view_migration', id=id))

@app.route('/migration/<int:id>/delete', methods=['POST'])
@login_required
def delete_migration(id):
    if current_user.team.name != 'SD' and not current_user.is_admin:
        flash('Only SD team members can delete migrations.', 'error')
        return redirect(url_for('index'))

    migration = Migration.query.get_or_404(id)
    
    # Check if user is the creator or admin
    if migration.created_by != current_user.id and not current_user.is_admin:
        flash('You can only delete your own migrations.', 'error')
        return redirect(url_for('index'))

    try:
        # Instead of deleting, mark as deleted
        migration.is_deleted = True
        migration.status = 'deleted'  # Optional: add a deleted status
        db.session.commit()

        flash('Migration deleted successfully.', 'success')
        log_action(migration.id, current_user.id, 'deleted', 'Migration marked as deleted')
        
    except Exception as e:
        db.session.rollback()
        flash('Error deleting migration.', 'error')
        print(f"Error: {str(e)}")  # For debugging

    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    # Get date range from query parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    # Create base query
    base_query = Migration.query.filter_by(is_deleted=False)

    # Apply date filtering if dates are provided
    if start_date and end_date:
        try:
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)  # Include the end date
            base_query = base_query.filter(Migration.created_at.between(start, end))
        except ValueError:
            flash('Invalid date format', 'error')

    # Get migration statistics with date filter
    total_migrations = base_query.count()
    completed = base_query.filter_by(status='completed').count()
    in_progress = base_query.filter_by(status='in_progress').count()
    rollback = base_query.filter_by(status='rollback').count()
    waiting = base_query.filter_by(status='waiting').count()
    acknowledged = base_query.filter_by(status='acknowledged').count()

    stats = {
        'total': total_migrations,
        'completed': completed,
        'in_progress': in_progress,
        'rollback': rollback,
        'waiting': waiting,
        'acknowledged': acknowledged
    }

    return render_template('dashboard.html', 
                         stats=stats, 
                         start_date=start_date, 
                         end_date=end_date)

@app.route('/admin/user-management')
@login_required
def user_manage():
    if not current_user.is_admin:
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('index'))
    
    users = User.query.all()
    teams = Team.query.all()
    return render_template('admin/user_manage.html', users=users, teams=teams)

@app.route('/admin/user/<int:user_id>/update', methods=['POST'])
@login_required
def update_user(user_id):
    if not current_user.is_admin:
        return jsonify({'status': 'error', 'message': 'Admin privileges required'})

    user = User.query.get_or_404(user_id)
    action = request.form.get('action')

    if action == 'toggle_active':
        user.is_active = not user.is_active
    elif action == 'toggle_admin':
        user.is_admin = not user.is_admin
    elif action == 'update_team':
        new_team_id = request.form.get('team_id')
        if new_team_id:
            user.team_id = new_team_id

    try:
        db.session.commit()
        return jsonify({'status': 'success'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)})

@app.errorhandler(404)
def page_not_found(e):
    return render_template('admin/404.html'), 404

@app.route('/migration/file/<int:file_id>/delete', methods=['POST'])
@login_required
def delete_migration_file(file_id):
    file = MigrationFile.query.get_or_404(file_id)
    migration = Migration.query.get(file.migration_id)
    
    # Check permissions
    if current_user.team.name != 'SD' and not current_user.is_admin:
        flash('Only SD team members can delete files.', 'error')
        return redirect(url_for('edit_migration_page', id=migration.id))
    
    if migration.created_by != current_user.id and not current_user.is_admin:
        flash('You can only delete files from your own migrations.', 'error')
        return redirect(url_for('edit_migration_page', id=migration.id))

    try:
        # Delete file from storage
        if os.path.exists(file.file_path):
            os.remove(file.file_path)
        
        # Delete from database
        db.session.delete(file)
        db.session.commit()
        
        flash('File deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting file.', 'error')
        print(f"Error: {str(e)}")  # For debugging
    
    return redirect(url_for('edit_migration_page', id=migration.id))

@app.route('/migration/file/<int:file_id>/replace', methods=['POST'])
@login_required
def replace_migration_file(file_id):
    file = MigrationFile.query.get_or_404(file_id)
    migration = Migration.query.get(file.migration_id)
    
    # Check permissions
    if current_user.team.name != 'SD' and not current_user.is_admin:
        flash('Only SD team members can replace files.', 'error')
        return redirect(url_for('edit_migration_page', id=migration.id))
    
    if migration.created_by != current_user.id and not current_user.is_admin:
        flash('You can only replace files from your own migrations.', 'error')
        return redirect(url_for('edit_migration_page', id=migration.id))

    if 'new_file' not in request.files:
        flash('No file uploaded.', 'error')
        return redirect(url_for('edit_migration_page', id=migration.id))
    
    new_file = request.files['new_file']
    if new_file.filename == '':
        flash('No file selected.', 'error')
        return redirect(url_for('edit_migration_page', id=migration.id))

    try:
        # Delete old file
        if os.path.exists(file.file_path):
            os.remove(file.file_path)
        
        # Save new file
        filename = secure_filename(new_file.filename)
        migration_folder = os.path.join(app.config['UPLOAD_FOLDER'], f'migration_{migration.id}')
        file_path = os.path.join(migration_folder, filename)
        new_file.save(file_path)
        
        # Update database record
        file.filename = filename
        file.file_path = file_path
        file.uploaded_at = datetime.utcnow()
        db.session.commit()
        
        flash('File replaced successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error replacing file.', 'error')
        print(f"Error: {str(e)}")  # For debugging
    
    return redirect(url_for('edit_migration_page', id=migration.id))

@app.route('/migration/<int:id>/files/delete', methods=['POST'])
@login_required
def delete_migration_files(id):
    migration = Migration.query.get_or_404(id)
    
    # Check permissions
    if current_user.team.name != 'SD' and not current_user.is_admin:
        flash('Only SD team members can delete files.', 'error')
        return redirect(url_for('edit_migration_page', id=id))
    
    if migration.created_by != current_user.id and not current_user.is_admin:
        flash('You can only delete files from your own migrations.', 'error')
        return redirect(url_for('edit_migration_page', id=id))

    file_ids = request.form.getlist('file_ids[]')
    
    if not file_ids:
        flash('No files selected for deletion.', 'error')
        return redirect(url_for('edit_migration_page', id=id))

    try:
        for file_id in file_ids:
            file = MigrationFile.query.get(file_id)
            if file and file.migration_id == id:  # Ensure file belongs to this migration
                # Delete file from storage
                if os.path.exists(file.file_path):
                    os.remove(file.file_path)
                
                # Delete from database
                db.session.delete(file)
        
        db.session.commit()
        flash('Selected files deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Error deleting files.', 'error')
        print(f"Error: {str(e)}")  # For debugging
    
    return redirect(url_for('edit_migration_page', id=id))

@app.errorhandler(500)
def internal_server_error(e):
    # Log the error details
    app.logger.error(f'Server Error: {e}')
    
    # In development, we might want to see the error details
    if current_app.config['DEBUG']:
        # Still show our custom 500 page but include error details
        error_details = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return render_template('admin/500.html', error_details=error_details), 500
    
    # In production, just show the generic error page
    return render_template('admin/500.html'), 500

# 403 Forbidden
@app.errorhandler(403)
def forbidden_error(e):
    app.logger.error(f'Forbidden: {e}')
    return render_template('admin/403.html'), 403


# 401 Unauthorized
@app.errorhandler(401)
def unauthorized_error(e):
    app.logger.error(f'Unauthorized: {e}')
    return render_template('admin/401.html'), 401

@app.route('/test-500')
def test_500():
    # Simulate a server error
    raise Exception("Test 500 error")

@app.route('/test-403')
def test_403():
    abort(403)

@app.route('/test-401')
def test_401():
    abort(401)

@app.route('/export-statistics')
@login_required
def export_statistics():
    # Get date range from query parameters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Create base query
    base_query = Migration.query.filter_by(is_deleted=False)
    
    # Apply date filter if provided
    if start_date and end_date:
        try:
            # Add time components to make the range inclusive
            start = datetime.strptime(start_date, '%Y-%m-%d')
            end = datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1)  # Include the end date
            base_query = base_query.filter(Migration.created_at.between(start, end))
        except ValueError:
            flash('Invalid date format', 'error')
    
    # Create workbook and select active sheet for SA Team Statistics
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "SA Team Statistics"
    
    # Create second sheet for Total Migrations
    ws2 = wb.create_sheet("Total Migrations")
    
    # Define styles
    header_fill = PatternFill(start_color="F1653F", end_color="F1653F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: SA Team Statistics
    headers1 = [
        "Team Member", 
        "Total Assigned", 
        "Completed", 
        "Rollback",
        "In Progress",
        "Waiting",
        "Acknowledged"
    ]
    
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Get SA team members
    sa_team = Team.query.filter_by(name='SA').first()
    sa_members = User.query.filter_by(team_id=sa_team.id).all()
    
    # Write data for each SA member
    row = 2
    for member in sa_members:
        # Get statistics for this member
        total = base_query.filter_by(assigned_to=member.id).count()
        completed = base_query.filter_by(assigned_to=member.id, status='completed').count()
        rollback = base_query.filter_by(assigned_to=member.id, status='rollback').count()
        in_progress = base_query.filter_by(assigned_to=member.id, status='in_progress').count()
        waiting = base_query.filter_by(assigned_to=member.id, status='waiting').count()
        acknowledged = base_query.filter_by(assigned_to=member.id, status='acknowledged').count()
        
        # Write member data
        ws1.cell(row=row, column=1, value=member.username).border = border
        ws1.cell(row=row, column=2, value=total).border = border
        ws1.cell(row=row, column=3, value=completed).border = border
        ws1.cell(row=row, column=4, value=rollback).border = border
        ws1.cell(row=row, column=5, value=in_progress).border = border
        ws1.cell(row=row, column=6, value=waiting).border = border
        ws1.cell(row=row, column=7, value=acknowledged).border = border
        
        row += 1
    
    # Add totals row
    ws1.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 8):
        ws1.cell(row=row, column=col, value=f"=SUM({chr(64+col)}2:{chr(64+col)}{row-1})").font = Font(bold=True)
    
    # Sheet 2: Total Migrations Statistics
    headers2 = [
        "Status",
        "Count",
        "Percentage"
    ]
    
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Get total statistics
    total_migrations = base_query.count()
    statuses = ['completed', 'rollback', 'in_progress', 'waiting', 'acknowledged']
    
    row = 2
    for status in statuses:
        count = base_query.filter_by(status=status).count()
        percentage = (count / total_migrations * 100) if total_migrations > 0 else 0
        
        # Write status data
        ws2.cell(row=row, column=1, value=status.title()).border = border
        ws2.cell(row=row, column=2, value=count).border = border
        ws2.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
        row += 1
    
    # Add total row
    ws2.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=row, column=2, value=total_migrations).font = Font(bold=True)
    ws2.cell(row=row, column=3, value="100%").font = Font(bold=True)
    
    # Add date range info to both sheets if provided
    if start_date and end_date:
        ws1.cell(row=row+2, column=1, value=f"Date Range: {start_date} to {end_date}")
        ws2.cell(row=row+2, column=1, value=f"Date Range: {start_date} to {end_date}")
    
    # Adjust column widths for both sheets
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename with date
    filename = f"migration_statistics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == "__main__":
    # Set debug to False to see the 500 error page
    app.config['DEBUG'] = True

    app.run(debug=True, host='0.0.0.0', port=5000)
    #socketio.run(app, host='0.0.0.0', port=5000)